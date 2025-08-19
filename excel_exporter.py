import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("警告：未安装 openpyxl。将无法生成格式化的 Excel 报表。")
    print("      可以通过运行 'pip install openpyxl' 来安装。")

from data_processor import COLUMN_NAME_MAP

def format_excel_sheet(sheet, df, title="", header_font=Font(bold=True), header_fill=PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")):
    """为 Excel 工作表应用基本格式"""
    if title:
        sheet.title = title

    # 创建df的副本并转换Period类型列
    df_processed = df.copy()
    for col in df_processed.columns:
        if isinstance(df_processed[col].dtype, pd.PeriodDtype):
            df_processed[col] = df_processed[col].astype(str)

    # 应用列名映射
    df_display = df_processed.rename(columns=COLUMN_NAME_MAP)

    for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column width (approximate)
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50) # Max width 50

def save_results_to_excel(account_month_summary, stock_summary, stock_detail_summary, details_df, output_file):
    """将结果保存到格式化的 Excel 文件"""
    # 在保存前，先将所有DataFrame中的Period类型列转换为字符串
    if account_month_summary is not None and not account_month_summary.empty:
        account_month_summary = account_month_summary.copy()
        for col in account_month_summary.columns:
            if isinstance(account_month_summary[col].dtype, pd.PeriodDtype):
                account_month_summary[col] = account_month_summary[col].astype(str)
    
    if stock_summary is not None and not stock_summary.empty:
        stock_summary = stock_summary.copy()
        for col in stock_summary.columns:
            if isinstance(stock_summary[col].dtype, pd.PeriodDtype):
                stock_summary[col] = stock_summary[col].astype(str)
    
    if stock_detail_summary is not None and not stock_detail_summary.empty:
        stock_detail_summary = stock_detail_summary.copy()
        for col in stock_detail_summary.columns:
            if isinstance(stock_detail_summary[col].dtype, pd.PeriodDtype):
                stock_detail_summary[col] = stock_detail_summary[col].astype(str)
    
    if details_df is not None and not details_df.empty:
        details_df = details_df.copy()
        for col in details_df.columns:
            if isinstance(details_df[col].dtype, pd.PeriodDtype):
                details_df[col] = details_df[col].astype(str)
    
    if not OPENPYXL_AVAILABLE:
        try:
            # 应用列名映射
            acc_mon_summary_display = account_month_summary.rename(columns=COLUMN_NAME_MAP)
            stock_summary_display = stock_summary.rename(columns=COLUMN_NAME_MAP)
            stock_det_summary_display = stock_detail_summary.rename(columns=COLUMN_NAME_MAP)
            details_df_display = details_df.rename(columns=COLUMN_NAME_MAP)
            # 交换买/卖时间列名以匹配显示逻辑
            if '买入时间' in details_df_display.columns and '卖出时间' in details_df_display.columns:
                details_df_display.rename(columns={'买入时间': 'temp_buy', '卖出时间': 'temp_sell'}, inplace=True)
                details_df_display.rename(columns={'temp_buy': '卖出时间', 'temp_sell': '买入时间'}, inplace=True)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                if not acc_mon_summary_display.empty:
                    acc_mon_summary_display.to_excel(writer, sheet_name='账户月度汇总', index=False)
                if not stock_summary_display.empty:
                    stock_summary_display.to_excel(writer, sheet_name='股票汇总', index=False)
                if not stock_det_summary_display.empty:
                    stock_det_summary_display.to_excel(writer, sheet_name='股票明细', index=False)
                if not details_df_display.empty:
                    # 格式化日期时间列
                    if '卖出时间' in details_df_display.columns:
                        details_df_display['卖出时间'] = pd.to_datetime(details_df_display['卖出时间']).dt.strftime('%Y-%m-%d %H:%M:%S')
                    if '买入时间' in details_df_display.columns:
                        details_df_display['买入时间'] = pd.to_datetime(details_df_display['买入时间']).dt.strftime('%Y-%m-%d %H:%M:%S')
                    # 转换Period类型的month列为字符串
                    if '月份' in details_df_display.columns:
                        details_df_display['月份'] = details_df_display['月份'].astype(str)
                    details_df_display.to_excel(writer, sheet_name='交易匹配明细', index=False)
            return True, f"结果已保存到 '{output_file}' (基础格式，建议安装 openpyxl 获得美化效果)"
        except Exception as e:
            return False, f"保存Excel文件时出错1: {e}"

    try:
        wb = Workbook()
        wb.remove(wb.active)

        if not account_month_summary.empty:
            ws1 = wb.create_sheet(title="账户月度汇总")
            format_excel_sheet(ws1, account_month_summary, "账户月度汇总")

        if not stock_summary.empty:
            ws_new = wb.create_sheet(title="股票汇总")
            format_excel_sheet(ws_new, stock_summary, "股票汇总")

        if not stock_detail_summary.empty:
            ws2 = wb.create_sheet(title="股票明细")
            format_excel_sheet(ws2, stock_detail_summary, "股票明细")

        if not details_df.empty:
            ws3 = wb.create_sheet(title="交易匹配明细")
            details_with_formatted_dates = details_df.copy()
            # 交换买/卖时间列名以匹配显示逻辑
            if 'sell_datetime' in details_with_formatted_dates.columns and 'buy_datetime' in details_with_formatted_dates.columns:
                details_with_formatted_dates.rename(columns={'sell_datetime': 'temp_buy', 'buy_datetime': 'temp_sell'}, inplace=True)
                details_with_formatted_dates.rename(columns={'temp_buy': 'buy_datetime', 'temp_sell': 'sell_datetime'}, inplace=True)
            
            if 'sell_datetime' in details_with_formatted_dates.columns:
                details_with_formatted_dates['sell_datetime'] = pd.to_datetime(details_with_formatted_dates['sell_datetime']).dt.strftime('%Y-%m-%d %H:%M:%S')
            if 'buy_datetime' in details_with_formatted_dates.columns:
                details_with_formatted_dates['buy_datetime'] = pd.to_datetime(details_with_formatted_dates['buy_datetime']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # 转换Period类型的month列为字符串
            if 'month' in details_with_formatted_dates.columns:
                details_with_formatted_dates['month'] = details_with_formatted_dates['month'].astype(str)
            
            format_excel_sheet(ws3, details_with_formatted_dates, "交易匹配明细")

        wb.save(output_file)
        return True, f"格式化的结果已保存到 '{output_file}'"
    except Exception as e:
        # 输出错误信息
        print(f"保存Excel文件时出错: {e}")
        return False, f"保存Excel文件时出错2: {e}"